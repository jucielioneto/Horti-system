from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def export_store_excel(rows: List[Dict[str, Any]], store_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = store_name[:31]

    headers = ["CODIGO DO PRODUTO", "NOME DO PRODUTO", "QUANTIDADE", "UNIDADE"]
    ws.append(headers)

    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="355E3B")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in rows:
        ws.append([
            row.get("code"),
            row.get("name"),
            row.get("quantity"),
            row.get("unit"),
        ])

    for column_cells in ws.columns:
        max_length = 12
        for cell in column_cells:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 50)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


