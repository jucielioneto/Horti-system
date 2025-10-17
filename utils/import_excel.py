import os
from typing import Optional
from openpyxl import load_workbook

from models.produto import upsert_product


def import_products_from_excel(path: str) -> int:
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    wb = load_workbook(path)
    ws = wb.active
    count = 0
    # Expect columns: code, name, unit somewhere in first row. Try to infer headers.
    headers = {str((ws.cell(row=1, column=c).value or '')).strip().lower(): c for c in range(1, ws.max_column+1)}
    # Common header names
    code_col = headers.get('codigo') or headers.get('código') or headers.get('code') or 1
    name_col = headers.get('nome') or headers.get('produto') or headers.get('name') or 2
    unit_col = headers.get('unidade') or headers.get('um') or headers.get('unit') or 3

    for r in range(2, ws.max_row+1):
        code = str(ws.cell(row=r, column=code_col).value or '').strip()
        name = str(ws.cell(row=r, column=name_col).value or '').strip()
        unit = str(ws.cell(row=r, column=unit_col).value or '').strip().upper()
        if not code or not name:
            continue
        if unit not in ('KG','UN'):
            # Try to normalize
            if unit.startswith('K'):
                unit = 'KG'
            else:
                unit = 'UN'
        upsert_product(code, name, unit)
        count += 1
    return count


